import os
from openpyxl import Workbook

def init_db():
    db_path = "Rent_Management.xlsx"
    if not os.path.exists(db_path):
        wb = Workbook()

        # Properties Sheet
        ws_prop = wb.active
        ws_prop.title = "Properties"
        ws_prop.append(["Prop_ID", "Title", "Address", "Rent", "Vacancy",
                         "WiFi", "Security", "Parking", "Furnished", "Kitchen", "AC", "Electricity", "Gas"])

        # Customers Sheet
        ws_cust = wb.create_sheet("Customers")
        ws_cust.append(["Cust_ID", "Name", "Contact", "Property_ID", "Duration", "Property_Type"])

        # Payments Sheet
        ws_pay = wb.create_sheet("Payments")
        ws_pay.append(["Pay_ID", "Cust_ID", "Name", "Amount_Paid", "Month", "Status"])

        # Electricity Sheet
        ws_el = wb.create_sheet("Electricity")
        ws_el.append(["Rec_ID", "Cust_ID", "Customer", "Property", "Month",
                       "Prev_Read", "Curr_Read", "Units", "Energy_Charge",
                       "Fuel_Adj", "Fixed", "GST", "Total_Bill", "KE_Acc", "Meter_No"])

        wb.save(db_path)
        print("Database initialized successfully.")
    else:
        print("Database already exists.")

if __name__ == "__main__":
    init_db()
